import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

move_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g3')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g3_moves"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/attackdex/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]').css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        s = response.xpath('(//td[@class="fooevo"])[7]/br/following-sibling::text()').get().strip().replace(' %','').replace('--','—')
        a = response.xpath('(//td[@class="cen"])[5]').get().split('.gif"><img')
        j = response.xpath('(//td[@class="cen"])[6]').get().split('.gif"><img')
        if s.isdigit():
            effect_rate = int(s)
        else:
            effect_rate = s
        if a != ['<td class="cen">\n\t\t</td>']:
            appeal = len(a)
        else:
            appeal = 0
        if j != ['<td class="cen">\n\t\t</td>']:
            jam = len(j)
        else:
            jam = 0

        entry = {
            'move_name': response.xpath('//td[@align="center"][1]/text()').get().strip(),
            'battle_type': response.xpath('//td[@align="center"][2]//img[starts-with(@src, "/attackdex/type/")]/@src').get().replace('/attackdex/type/','').replace('.gif','').capitalize(),
            'power_points': int(response.xpath('(//td[@class="cen"])[1]//text()').get().strip()),
            'base_power': int(response.xpath('(//td[@class="cen"])[2]//text()').get().strip()),
            'accuracy': int(response.xpath('(//td[@class="cen"])[3]//text()').get().strip()),
            'battle_effect': response.xpath('(//td[@class="fooinfo"])[1]/text()').get().strip(),
            'secondary_effect': response.xpath('(//td[@class="fooinfo"])[2]/text()').get().strip(),
            'effect_rate': effect_rate,
            'contest_type': response.xpath('//td[@align="center"][3]//img[starts-with(@src, "/attackdex/type/")]/@src').get().replace('/attackdex/type/','').replace('.gif','').capitalize(),
            'contest_effect': response.xpath('(//td[@class="fooinfo"])[3]/text()').get().strip(),
            'appeal': appeal,
            'jam': jam,
            'tm_no': response.xpath('(//td[@class="cen"])[4]//text()').get().strip().replace('None',''),
            'bright_powder': response.xpath('(//td[@class="cen"])[7]//text()').get().strip(),
            'kings_rock': response.xpath('(//td[@class="cen"])[8]//text()').get().strip(),
            'physical_contact': response.xpath('(//td[@class="cen"])[9]//text()').get().strip(),
            'battle_palace_style': response.xpath('(//td[@class="cen"])[10]//text()').get().strip(),
            'battle_arena': response.xpath('(//td[@class="cen"])[11]//text()').get().strip(),
            'pokemon_hit': response.xpath('(//td[@class="cen"])[12]//text()').get().strip()
        }

        move_data.append(entry)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''

        # create datafrane from the global list 'move_data'
        df = pd.DataFrame(move_data)
        df = df.sort_values(by=['move_name'])

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g3_moves.xlsx')
        df.to_sql('moves', engine, index=False)
        #df.to_csv('pkmn_g3_moves.csv', index=False)
