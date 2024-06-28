import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

move_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g4')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g4_moves"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/attackdex-dp/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]').css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        f = response.xpath('(//td[@class="fooleft"])[2]/text()').get()
        s = response.xpath('//td[text()="Effect Rate:"]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"]/text()').get().strip().replace(' %','').replace('--','—')
        a = response.xpath('//td[b[text()="Points:"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"]').get().split('.gif"><img')
        if f == 'In-Depth Effect:':
            further_info = 'Yes'
        else:
            further_info = 'No'
        if s.isdigit():
            effect_rate = int(s)
        else:
            effect_rate = s
        if a != ['<td class="cen">\r\n\xa0</td>']:
            points = len(a)
        else:
            points = 0

        entry = {
            'move_name': response.xpath('(//td[@class="cen"])[1]/text()').get().strip(),
            'battle_type': response.xpath('(//td[@class="cen"])[2]//img[starts-with(@src, "/attackdex/type/")]/@src').get().replace('/attackdex/type/','').replace('.gif','').capitalize(),
            'category': response.xpath('(//td[@class="cen"])[3]//img[starts-with(@src, "/pokedex-dp/type/")]/@src').get().replace('/pokedex-dp/type/','').replace('.png','').capitalize(),
            'power_points': int(response.xpath('(//td[@class="cen"])[4]//text()').get().strip()),
            'base_power': int(response.xpath('(//td[@class="cen"])[5]//text()').get().strip()),
            'accuracy': int(response.xpath('(//td[@class="cen"])[6]//text()').get().strip()),
            'battle_effect': response.xpath('(//td[@class="fooinfo"])[1]/text()').get().strip(),
            'further_info': further_info,
            'secondary_effect': response.xpath('//td[text()="Secondary Effect:"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"]/text()').get().strip(),
            'effect_rate': effect_rate,
            'contest_type': response.xpath('//td[text()="Super Contests"]/following-sibling::td[@class="fooinfo"]//img[starts-with(@src, "/attackdex-dp/type/")]/@src').get().replace('/attackdex-dp/type/','').replace('.gif','').capitalize(),
            'contest_effect': response.xpath('//td[b[text()="Effect:"]]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"]/text()').get().strip(),
            'contest_points': points,
            'tm_no': response.xpath('//td[b[text()="TM #"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"][1]/text()').get().strip(),
            'speed_priority': response.xpath('//td[b[text()="Speed Priority"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"][2]/text()').get().strip(),
            'pokemon_hit': response.xpath('//td[b[text()="Pokémon Hit in Battle"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"][3]/text()').get().strip(),
            'bright_powder': response.xpath('//td[b[text()="BrightPowder"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"][1]/text()').get().strip(),
            'kings_rock': response.xpath('//td[b[text()="King\'s Rock"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"][2]/text()').get().strip(),
            'physical_contact': response.xpath('//td[b[text()="Physical Contact"]]/ancestor::tr/following-sibling::tr[1]//td[@class="cen"][3]/text()').get().strip()
        }

        move_data.append(entry)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''

        # create datafrane from the global list 'move_data'
        df = pd.DataFrame(move_data)
        df = df.sort_values(by=['move_name'])

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g4_moves.xlsx')
        df.to_sql('moves', engine, index=False)
        #df.to_csv('pkmn_g4_moves.csv', index=False)
