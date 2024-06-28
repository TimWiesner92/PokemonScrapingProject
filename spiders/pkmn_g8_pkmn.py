import scrapy, pandas as pd, math, re
from openpyxl import load_workbook
from sqlalchemy import create_engine

pk_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g8')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g8_pkmn"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex-swsh/"]

    def parse(self, response):
        '''Method for extracting links to pages to crawl from the pulldowns'''
        pulldown_links = [x for x in response.xpath('//select[@name="SelectURL"]')[0:8].css('option').xpath('@value').getall() if x != '#']

        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        '''Method for pre-processing and allocating data pulled from each page'''
        # define functions used in processing data
        def reduce_fractions(num, denom):
            '''Function to reduce a fraction using greatest common demoninator. E.g., reduce '5/25' down to '1/5'. Takes num and denom, reduces each, and returns reduced_num and reduced_denom'''
            gcd = math.gcd(int(num), int(denom))
            reduced_num = int(num) // gcd
            reduced_denom = int(denom) // gcd
            return reduced_num, reduced_denom
        
        def extract_types(html_string, pat1, pat2, pk_name):
            '''Function to extract text from an html string based on a pre-defined pattern'''
            matches = pat1.findall(html_string)
            types_match = [match.capitalize() for match in matches]
            form_match = pat2.search(html_string).group(1)
            # Darmanitan's forms are named differently in Type than they are in Alternate Forms
            if pk_name == 'Darmanitan':
                if form_match == 'Normal':
                    form_match = 'Normal Standard Mode'
                elif form_match == 'Zen Mode':
                    if types_match == ['Fire', 'Psychic']:
                        form_match = 'Normal Zen Mode'
                    else:
                        form_match = 'Galarian Form Zen Mode'
                else:
                    form_match = 'Galarian Form Standard Mode'
            elif pk_name == 'Necrozma' and form_match == 'Normal':
                form_match = 'Necrozma'
            elif form_match == 'Alolan':
                form_match = 'Alola Form'
            elif form_match == 'Galarian':
                form_match = 'Galarian Form'
            elif form_match == 'Hisuian':
                form_match = 'Hisuian Form'
            else:
                form_match = form_match
            return (form_match,types_match)
        
        def ev_split(html):
            def list_split(ev_list):
                evs = [0,0,0,0,0,0]
                for i in ev_list:
                    if i[2:4] == 'HP':
                        evs[0] = int(i[0])
                    elif i[2:8] == 'Attack':
                        evs[1] = int(i[0])
                    elif i[2:9] == 'Defense':
                        evs[2] = int(i[0])
                    elif i[2:12] == 'Sp. Attack':
                        evs[3] = int(i[0])
                    elif i[2:13] == 'Sp. Defense':
                        evs[4] = int(i[0])
                    else:
                        evs[5] = int(i[0])
                return evs

            if len(html) == 1:
                html_split = html[0].split('<br>')
                ev_iter = list_split(html_split)
            else:
                forms = []
                points = []
                for i in html:
                    form, point = i.split('</b><br>')
                    forms.append(form)
                    points.append(point)
                forms = [i.replace('<b>','') for i in forms]
                points = [i.split('<br>') for i in points]
                ev_list_split = [list_split(i) for i in points]
                ev_iter = {forms[i]: ev_list_split[i] for i in range(len(html))}
            return ev_iter

        # converts raw data on gender into 'N/A' for genderless pokemon, 'Always female' or 'Always male' for those pokemon,
        # or produces a simplified ratio (e.g. '7:1' for startes) for pokemon that can be male or female
        if response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').get()[-19:-5] == 'Gender Unknown':
            mfr = 'N/A'
        else:
            mr = response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').css('td::text').getall()[2].replace('%','').replace('*','')
            fr = response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').css('td::text').getall()[5].replace('%','').replace('*','')
            if mr == '0' and fr == '0':
                mfr = 'double_zero (set back to N/A if this shows up, remove this clause if not)'
            elif mr == '0':
                mfr = 'Always female'
            elif fr == '0':
                mfr = 'Always male'
            else:
                reduced_mr, reduced_fr = reduce_fractions(int(100 * float(mr)), int(100 * float(fr)))
                mfr = str(reduced_mr) + ':' + str(reduced_fr)

        # splits awarded EVs for beating a pokemon into individual values so they each can be placed in their own columns for easier sorting and filtering
        evs_html = response.xpath('//td[text()="Effort Values Earned"]/ancestor::tr/following-sibling::tr[1]//td[3]').get().replace('<td class="fooinfo">','').replace('</td>','')[:-4]
        if 'Zen Mode' in evs_html:
            evs_html = evs_html
        elif 'Alola Form' in evs_html:
            evs_html = 'Normal</b><br>' + evs_html
        elif 'Galarian Form' in evs_html:
            evs_html = 'Normal</b><br>' + evs_html
        elif 'Hisuian Form' in evs_html:
            evs_html = 'Normal</b><br>' + evs_html
        else:
            evs_html = evs_html
        evs_html = evs_html.split('<br><b>')

        # height, abilities, and egg_groups declared here for simpler list comprehensions
        # splits height into feet and inches so height can be transformed into a float
        if response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get() == '\r\n\t':
            base_height_raw = response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]//td[text()="Standard"]/following-sibling::td[1]/text()').get().split(' / ')
            base_weight = [float(i) for i in response.xpath('//td[text()="Weight"]/ancestor::tr/following-sibling::tr[1]//td[3]//td[text()="Standard"]/following-sibling::td[1]/text()').get().replace(',','').replace('lbs','').split(' / ')]
        else:
            base_height_raw = response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().split(' / ')
            base_weight = [float(i) for i in response.xpath('//td[text()="Weight"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().replace(',','').replace('lbs','').split(' / ')]
        base_height_sliced = [i.replace('"','').replace(' / ','\'').split('\'') for i in base_height_raw]
        base_height = [round((float(i[0]) + (float(i[1]) / 12)),2) for i in base_height_sliced]
        
        # splits abilities into individual values to put in separate columns for easier filtering
        abilities_raw = ''.join(response.xpath('//td[b/text()="Abilities"]//text()').getall()).strip().split(' Abilities: ')
        abilities = [abilities_raw[i].replace('Abilities: ','').replace('- B2W2','B2W2').split(' - ') for i in range(len(abilities_raw))]
        egg_groups_selector = response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"]//a//text()').getall()
        egg_steps_selector = response.xpath('//td[text()="Base Egg Steps"]/ancestor::tr/following-sibling::tr[1]//td[5]/text()').get().replace(',','').replace(' (SWSH)','')
        if response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"]//text()').get()[-12:] == 'cannot breed':
            eg1,eg2 = ['No Eggs Discovered','']
        elif len(egg_groups_selector) == 1:
            eg1,eg2 = [egg_groups_selector[0],'']
        else:
            eg1 = egg_groups_selector[0]
            eg2 = egg_groups_selector[1]
        
        # set up types_html, stats_forms, and name for if/else block
        types_html = response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//tr').getall()
        stats_forms = response.xpath('//table[@class="dextable"][.//td[h2[starts-with(text(), "Stats")]]]//td[@class="fooinfo" and contains(text(), "Base Stats")]').getall()
        stats_list = [[int(i) for i in response.xpath('(//table[@class="dextable"][.//td[h2[starts-with(text(), "Stats")]]])[' + str(j + 1) + ']//td[@class="fooinfo" and contains(text(), "Base Stats")]/following-sibling::td/text()').getall()] for j in range(len(stats_forms))]
        name = response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get()
        cptr_rt = response.xpath('//td[text()="Capture Rate"]/ancestor::tr/following-sibling::tr[1]//td[4]/text()').get()

        # set up forms, types, etc. to be parsed based on whether a pokemon has multiple distinct forms
        mega_primal_forms = list(set(response.xpath('//td[@class="fooevo" and @colspan="6"]//h2//text()').getall()))
        p1 = re.compile(r'/pokedex-bw/type/([a-z]+)\.gif')
        p2 = re.compile(r'<td width="50%">(.*?)</td>')

        if len(mega_primal_forms) > 0:
            forms = [''] + mega_primal_forms
            base_types = [[i.replace('/pokedex-bw/type/','').replace('.gif','').capitalize() for i in response.xpath('(//td[text()="Picture"]/ancestor::table/following-sibling::table)[1]//tr//td[5]//img[starts-with(@src, "/pokedex-bw/type/")]/@src').getall()]]
            # type, height, weight, and ability can all change upon mega evolving
            abilities = [ability for ability, count in zip(abilities, [3,1]) for _ in range(count)]
            mp_types = [[i.replace('/pokedex-bw/type/','').replace('.gif','').capitalize() for i in response.xpath('//table[@class="dextable"]//td/h2[text()="' + mega_primal_forms[0] + '"]/ancestor::table/following::table[1]//tr[2]//td[5]//img[starts-with(@src, "/pokedex-bw/type/")]/@src').getall()]]
            mp_height_raw = response.xpath('//table[@class="dextable"]//td/h2[text()="' + mega_primal_forms[0] + '"]/ancestor::table/following::table[1]//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().replace('"','').replace(' / ','\'').split('\'')
            mp_height = [round((float(mp_height_raw[0]) + (float(mp_height_raw[1]) / 12)),2)]
            mp_weight = [response.xpath('//table[@class="dextable"]//td/h2[text()="' + mega_primal_forms[0] + '"]/ancestor::table/following::table[1]//td[text()="Weight"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().replace(',','').replace('lbs','')]
            mp_stats = [[int(i) for i in response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])//td[@class="fooinfo" and contains(text(), "Base Stats")]/following-sibling::td/text()').getall()]]
            types_match = base_types + mp_types
            heights_match = base_height + mp_height
            weights_match = base_weight + mp_weight
            types = {forms[i]: types_match[i] for i in range(len(forms))}
            heights = {forms[i]: heights_match[i] for i in range(len(forms))}
            weights = {forms[i]: weights_match[i] for i in range(len(forms))}
            stats_list += mp_stats
        elif len(response.xpath('//td[h2[contains(text(), "Stats")]]').getall()) > 1 or (len(types_html) > 1 and name != 'Castform'):
            forms = response.xpath('//td[text()="Alternate Forms"]/ancestor::tr/following-sibling::tr[1]//td/b/text()').getall()
            forms = [i.replace('Kantonian Form','Normal').replace('Johtonian Form','Normal').replace('Hoennian Form','Normal').replace('Unovan Form','Normal').replace('Kalosian Form','Normal').replace('Alolan Form','Normal') for i in forms]
            if forms == []:
                forms = ['']
            if name == 'Minior':
                forms = ['Meteor Form','Core Form']
                cptr_rt = [re.sub(r'\s*\(.*?\)', '', i) for i in response.xpath('//td[text()="Capture Rate"]/ancestor::tr/following-sibling::tr[1]//td[4]/text()').getall()]
            elif name == 'Darmanitan':
                forms = ['Normal Standard Mode','Normal Zen Mode','Galarian Form Standard Mode','Galarian Form Zen Mode']
            elif name in ['Indeedee','Basculegion']:
                forms = ['Male','Female']
            elif name in ['Pikachu','Burmy','Cherrim','Shellos','Gastrodon','Arceus']:
                forms = ['']
            if len(types_html) != 0:
                types = {form_match: types_match for form_match, types_match in (extract_types(html,p1,p2,name) for html in types_html)}
            else:
                types = [i.replace('/pokedex-bw/type/','').replace('.gif','').capitalize() for i in response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//img[starts-with(@src, "/pokedex-bw/type/")]/@src').getall()]
        else:
            forms = ['']
            types = [i.replace('/pokedex-bw/type/','').replace('.gif','').capitalize() for i in response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//img[starts-with(@src, "/pokedex-bw/type/")]/@src').getall()]

        
        for form in forms:
            # set up selectors for stats and types based on how a given pokemon's forms vary from one another
            if len(mega_primal_forms) > 0:
                height_sel = heights[form]
                weight_sel = weights[form]
                type_sel = types[form]
                abilities_sel = abilities[forms.index(form)]
                stat_sel = forms.index(form)
            else:
                height_sel = base_height[0]
                weight_sel = base_weight[0]
                abilities_sel = abilities[0]
                if name == 'Rotom':
                    type_sel = types[form]
                    if forms.index(form) == 0:
                        stat_sel = 0
                    else:
                        stat_sel = 1
                elif len(types_html) == len(forms) == len(stats_forms):
                    type_sel = types[form]
                    stat_sel = forms.index(form)
                elif name == 'Zygarde':
                    type_sel = types
                    if forms.index(form) == 0:
                        stat_sel = 1
                    elif forms.index(form) == 1:
                        stat_sel = 0
                    else:
                        stat_sel = 2
                elif name == 'Palkia' or name == 'Dialga':
                    type_sel = types
                    if forms.index(form) == 0:
                        stat_sel = 0
                    else:
                        stat_sel = 4
                elif name == 'Basculegion' or name == 'Enamorus':
                    type_sel = types
                    if forms.index(form) == 0:
                        stat_sel = 0
                    else:
                        stat_sel = 2
                elif name == 'Wormadam':
                    type_sel = types[form]
                    stat_sel = forms.index(form)
                elif len(types_html) == len(forms) and len(stats_forms) == 1:
                    type_sel = types[form]
                    stat_sel = 0
                elif len(types_html) == 0 and len(forms) == len(stats_forms):
                    type_sel = types
                    stat_sel = forms.index(form)
                elif len(types_html) == 0 and len(forms) != len(stats_forms):
                    type_sel = types
                    if forms.index(form) == 0:
                        stat_sel = 0
                    else:
                        stat_sel = 1
                elif len(types_html) != len(forms) == len(stats_forms):
                    type_sel = types[forms.index(form)]
                    if forms.index(form) == 0:
                        stat_sel = 0
                    else:
                        stat_sel = 1
                else:
                    type_sel = types[form]
                    if forms.index(form) == 0:
                        stat_sel = 0
                    else:
                        stat_sel = 1
            
            # set up selectors for evs:
            if len(evs_html) == 1:
                if evs_html == ['']:
                    evs_sel = ['Not listed','Not listed','Not listed','Not listed','Not listed','Not listed']
                else:
                    evs_sel = ev_split(evs_html)
            else:
                if name == 'Darmanitan':
                    if form == 'Normal Standard Mode':
                        evs_sel = ev_split(evs_html)['Standard']
                    elif form == 'Galarian Form Standard Mode':
                        evs_sel = ev_split(evs_html)['Galarian Form']
                    else:
                        evs_sel = ev_split(evs_html)['Zen Mode']
                else:
                    evs_sel = ev_split(evs_html)[form]
            
            # base_stats:
            base_stats = stats_list[stat_sel]
            
            entry = {
                'dex_no': response.xpath('//td[b[text()="National"]]/following-sibling::td[1]/text()').get().replace('#','').zfill(4),
                'form_index': forms.index(form),
                'name': name,
                'form': form,
                'class': response.xpath('//td[text()="Classification"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().replace(' Pokémon',''),
                'height_ft': height_sel,
                'weight_lb': weight_sel,
                'male_female_rate': mfr,
                'type_one': type_sel[0],
                'type_two': type_sel[1] if len(type_sel) == 2 else '',
                'ability_one': abilities_sel,
#                'ability_two': abilities[1] if len(abilities) == 2 else '',
                'base_hp': int(base_stats[0]),
                'base_atk': int(base_stats[1]),
                'base_def': int(base_stats[2]),
                'base_sp_atk': int(base_stats[3]),
                'base_sp_def': int(base_stats[4]),
                'base_speed': int(base_stats[5]),
                'cptr_rt': int(cptr_rt) if type(cptr_rt) == str else int(cptr_rt[forms.index(form)]),
                'exp_growth': response.xpath('//td[text()="Experience Growth"]/ancestor::tr/following-sibling::tr[1]//td[1]/br/following-sibling::text()').get(),
                'base_egg_steps': 'N/A' if egg_steps_selector == '\xa0' else int(egg_steps_selector),
                'egg_group_one': eg1,
                'egg_group_two': eg2,
                'dynamax': response.xpath('//td[text()="Dynamax Capable?"]/ancestor::tr/following-sibling::tr[1]//td[4]//text()').get().strip()[-11:].replace('can Dynamax','Yes').replace('not Dynamax','No'),
                'hp_evs': evs_sel[0],
                'atk_evs': evs_sel[1],
                'def_evs': evs_sel[2],
                'sp_atk_evs': evs_sel[3],
                'sp_def_evs': evs_sel[4],
                'speed_evs': evs_sel[5]
            }

            pk_data.append(entry)
        
    def closed(self, data):
        '''Method for post-processing and saving the data after the spider finishes'''
        # define functions used in this method
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        # create datafrane from the global list 'pk_data'
        df = pd.DataFrame(pk_data)
        df = df.sort_values(by=['dex_no','form_index'])

        # add suffixes to pokemon with alternate forms
        df['dex_no'] = df['dex_no'] + df['form'].str[:2].str.lower()

        # remove suffixes from Pokemon with 'standard' forms
        for i in ['0386no','0479ro','0555st','0646ky','0800ne','0898ca']:
            df.loc[df['dex_no'] == i, 'dex_no'] = df['dex_no'].str[0:4]

        # otherwise correct suffixes
        df.loc[df['form'] == 'Normal', 'dex_no'] = df['dex_no'].str[0:4]
        df.loc[df['form'] == 'Mega Evolution', 'dex_no'] = df['dex_no'].str[0:4] + 'meg'
        df.loc[df['form'] == 'Mega Evolution X', 'dex_no'] = df['dex_no'].str[0:4] + 'meg-x'
        df.loc[df['form'] == 'Mega Evolution Y', 'dex_no'] = df['dex_no'].str[0:4] + 'meg-y'
        df.loc[df['form'] == 'Normal Standard Mode', 'dex_no'] = df['dex_no'].str[0:4]
        df.loc[df['form'] == 'Normal Zen Mode', 'dex_no'] = df['dex_no'].str[0:4] + 'ze'
        df.loc[df['form'] == 'Galarian Form Standard Mode', 'dex_no'] = df['dex_no'].str[0:4] + 'g'
        df.loc[df['form'] == 'Galarian Form Zen Mode', 'dex_no'] = df['dex_no'].str[0:4] + 'g-ze'
        df.loc[df['form'] == 'Midday Form', 'dex_no'] = df['dex_no'].str[0:4] + 'md'
        df.loc[df['form'] == 'Midnight Form', 'dex_no'] = df['dex_no'].str[0:4] + 'mn'
        df.loc[df['form'] == 'Hoopa Confined', 'dex_no'] = df['dex_no'].str[0:4] + 'co'
        df.loc[df['form'] == 'Hoopa Unbound', 'dex_no'] = df['dex_no'].str[0:4] + 'un'
        df.loc[df['form'] == '10% Forme', 'dex_no'] = df['dex_no'].str[0:4] + 'te'
        df.loc[df['form'] == '50% Forme', 'dex_no'] = df['dex_no'].str[0:4] + 'fi'
        df.loc[df['form'] == 'Noice Face', 'form'] = 'NoIce Face'
        df.loc[df['form'] == 'Alola Form', 'form'] = 'Alolan Form'
        df.loc[df['form'] == 'Alolan Form', 'dex_no'] = df['dex_no'].str[0:4] + 'a'
        df.loc[df['form'] == 'Galarian Form', 'dex_no'] = df['dex_no'].str[0:4] + 'g'
        df.loc[df['form'] == 'Hisuian Form', 'dex_no'] = df['dex_no'].str[0:4] + 'h'
        df.loc[df['form'] == 'Male', 'male_female_rate'] = 'Always male'
        df.loc[df['form'] == 'Female', 'male_female_rate'] = 'Always female'

        # drop Unovan Samurott
        samurott = df[(df['dex_no'] == '0503')].index
        df = df.drop(samurott)

        # drop form_index
        df = df.drop(['form_index'], axis=1)

        # save to xlsx, PostgreSQL, or csv
        save_xlsx('pkmn_g8_pkmn.xlsx')
        df.to_sql('pkmn', engine, index=False)
        #df.to_csv('pkmn_g8_pkmn.csv', index=False)
