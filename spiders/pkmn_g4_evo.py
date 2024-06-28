import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

evo_data = []
methods = ['level','stone','trade','friendship','status']
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g4')

class PkmnSpider(scrapy.Spider):
    name = 'pkmn_g4_evo'

    def start_requests(self):
        '''The list of pages to crawl'''
        
        for i in methods:
            next_page='https://pokemondb.net/evolution/' + i
            yield scrapy.Request(next_page, callback=self.parse, meta={'page_type': i})

    def parse(self, response):
        '''Pull the info from each page and create a dictionary from the results'''
        for row in response.css('tbody tr'):
            evolves_from = row.css('td:nth-child(1) a::attr(title)').get().replace('View Pokedex for #','')[:4]
            evolves_from_form = row.css('td:nth-child(1) small.text-muted::text').get()
            evolves_to = row.css('td:nth-child(3) a::attr(title)').get().replace('View Pokedex for #','')[:4]
            evolves_to_form = row.css('td:nth-child(3) small.text-muted::text').get()
            method = response.meta['page_type']

            if method == 'stone':
                item_used = row.css('td.cell-med-text').xpath('string()').get().strip().replace(', outside Alola','').replace(' in Gen 8','').replace('+','')
                item_held = ''
                other_condition = ''
            elif method == 'trade':
                item_used = ''
                item_held = row.css('td.cell-med-text').xpath('string()').get().strip().replace('Kings Rock',"King's Rock")
                other_condition = ''
            else:
                item_used = ''
                item_held = ''
                other_condition = row.css('td.cell-med-text').xpath('string()').get().strip().replace('outside Alola','').replace('outside Galar','').replace('or Level 17 in Legends: Arceus','')

            def regional_forms_check(var, var_form):
                '''Checks if the evolution involves a regional form and adds appropriate suffixes to Pokedex numbers while rewriting the form for easier filtering'''
                if var_form is not None:
                    if var_form[0:6] == 'Alolan':
                        var += 'a'
                    elif var_form[0:8] == 'Galarian':
                        var += 'g'
                    elif var_form[0:7] == 'Hisuian':
                        var += 'h'
                    elif var_form[0:7] == 'Paldean':
                        var += 'p'
                    else:
                        var += var_form[0:2].lower()
                
                return var
            
            pk = regional_forms_check(evolves_from, evolves_from_form)
            et = regional_forms_check(evolves_to, evolves_to_form)
            
            entry = {
                'evolves_from': pk,
                'evolves_to': et,
                'method': method,
                'level': row.css('td.cell-num::text').get(),
                'item_used': item_used,
                'item_held': item_held,
                'stat': '',
                'time_of_day': '',
                'location': '',
                'move_learned': '',
                'gender': '',
                'in_party': '',
                'other_condition': other_condition,
                'notes': ''
            }
        
            evo_data.append(entry)

    def closed(self, data):
        '''Convert data to pandas DataFrame, clean and filter per gen, and save to Excel file'''

        # create datafrane from the global list 'evo_data'
        df = pd.DataFrame(evo_data)

        # pre-sort cleaning, accounting for whether or not different forms need to appear as separate entries
        # Burmy's forms only differ in how they evolve, but forms are mutable within the space of a battle
        df.loc[df['evolves_from'].str.startswith('0412'), 'evolves_from'] = '0412'
        df.loc[df['evolves_to'].str.startswith('0413'), 'notes'] = "Wormadam will take on Burmy's form at time of evolution, which can change based on the location of its most recent battle, but only if it actually leaves the ball."

        # sort by method, then evolves_to, then evolves_from
        df['method_index'] = df['method'].apply(lambda x: methods.index(x))
        df = df.sort_values(by=['evolves_from','evolves_to','method_index'])
        df = df.drop('method_index', axis=1)

        # remove duplicate entries
        duplicate_mask = df.duplicated(subset=['evolves_from','evolves_to'], keep='first')
        df = df[~duplicate_mask]

        # post-sort cleaning, moving evolution conditions into the appropriate columns
        def condition_mover(search_str,correct_col,correct_str):
            '''Locates rows containing a given string in item_used or other_condition (or both), removes it from those columns, and adds a replacement string in the target column.'''
            df[correct_col] = df.apply(lambda row: correct_str if (search_str in row['item_used'] or search_str in row['other_condition']) else row[correct_col], axis=1)
            df['item_used'] = df['item_used'].str.replace(search_str, '')
            df['other_condition'] = df['other_condition'].str.replace(search_str, '')
        
        # evolution cases requiring special handling before condition_mover() is run
        df.loc[df['evolves_to'] == '0350', ['method','item_held','stat']] = ['level','','Beauty']

        # evolutions with conditions in the wrong column
        condition_mover('Daytime','time_of_day','Daytime')
        condition_mover('Nighttime','time_of_day','Nighttime')
        condition_mover('level up in a Magnetic Field area','location','Mt Coronet (any section)')
        condition_mover('level up near an Icy Rock','location','Route 217 (northeast)')
        condition_mover('level up near a Mossy Rock','location','Eterna Forest (southwest)')
        condition_mover('in buildings','location','In buildings')
        condition_mover('in caves','location','In caves')
        condition_mover('in grass','location','In grass')
        condition_mover('Female','gender','Female')
        condition_mover('Male','gender','Male')

        # evolution cases requiring special handling after condition_mover() is run
        df.loc[df['location'] == 'Mt Coronet (any section)', ['method','item_used']] = ['level','']
        df.loc[df['location'] == 'Route 217 (northeast)', ['method','item_used']] = ['level','']
        df.loc[df['location'] == 'Eterna Forest (southwest)', ['method','item_used']] = ['level','']

        # remove leading and trailing commas and spaces
        df.loc[df['item_used'].str.endswith(', '), 'item_used'] = df['item_used'].str.rstrip(', ')
        df.loc[df['item_used'].str.endswith(','), 'item_used'] = df['item_used'].str.rstrip(',')
        df.loc[df['other_condition'].str.startswith(', '), 'other_condition'] = df['other_condition'].str[2:]
        df.loc[df['other_condition'].str.endswith(', '), 'other_condition'] = df['other_condition'].str.rstrip(', ')

        # move from other_condition to move_learned
        move_condition = df['other_condition'].str.endswith('learned')
        df.loc[move_condition, 'move_learned'] = df.loc[move_condition, 'other_condition'].str[6:-8]
        df.loc[move_condition, 'method'] = 'level'
        df.loc[move_condition, 'other_condition'] = ''

        # move from other_condition to item_held
        item_condition = df['other_condition'].str.startswith('hold ')
        df.loc[item_condition, 'item_held'] = df.loc[item_condition, 'other_condition'].str[5:]
        df.loc[item_condition, 'method'] = 'level'
        df.loc[item_condition, 'other_condition'] = ''

        # move from other_condition to stat
        stat_condition = df['other_condition'].str.startswith('Attack')
        df.loc[stat_condition, 'stat'] = df.loc[stat_condition, 'other_condition'].str[:]
        df.loc[stat_condition, 'other_condition'] = ''

        # move from other_condition to in_party
        df.loc[df['evolves_to'] == '0292', ['in_party','other_condition']] = ['Empty spot','Pokéball in bag']
        df.loc[df['evolves_to'] == '0226', ['method','in_party','other_condition']] = ['level','Remoraid','']

        # methods need to be renamed for better accuracy
        df.loc[df['method'] == 'stone', 'method'] = 'item'
        df.loc[df['method'] == 'friendship', ['method','stat']] = ['level','Friendship']

        # convert level to int, empty cells to NaN
        df['level'] = pd.to_numeric(df['level'], errors='coerce')
        df = df.replace('', pd.NA)

        # capitalize the first letter of other_condition
        df['other_condition'] = df['other_condition'].str[0].str.upper() + df['other_condition'].str[1:]

        # filter out entries not in the associated generation
        df = df[df['evolves_from'].between('0001', '0493') & ~df['evolves_from'].str.match(r'\d{4}[a-z]$')]
        df = df[df['evolves_to'].between('0001', '0493') & ~df['evolves_to'].str.match(r'\d{4}[a-z]$')]

        # drop any empty columns
        df = df.dropna(axis=1,how='all')

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g4_evo.xlsx')
        df.to_sql('evo', engine, index=False)
        #df.to_csv('pkmn_g4_evo.csv', index=False)
