import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

evo_data = []
methods = ['level','stone','trade','friendship','status']
games = ['','Brilliant Diamond, Shining Pearl','Sword, Shield, Brilliant Diamond, Shining Pearl','Legends: Arceus','Brilliant Diamond, Shining Pearl, Legends: Arceus','Sword, Shield','Sword','Shield','Pokémon GO']
unavailable_mons = ['0020a','0075a','0076a','0089a','0496','0497','0499','0500','0503','0505','0512','0514','0516','0523','0541','0542','0581','0586','0603','0604','0651','0652','0654','0655','0657','0658','0665','0666','0668','0670','0671','0673','0732','0733','0735','0740']
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g8')

class PkmnSpider(scrapy.Spider):
    name = 'pkmn_g8_evo'

    def start_requests(self):
        '''The list of pages to crawl'''
        
        for i in methods:
            next_page='https://pokemondb.net/evolution/' + i
            yield scrapy.Request(next_page, callback=self.parse, meta={'page_type': i})

    def parse(self, response):
        '''Pull the info from each page and create a dictionary from the results'''
        for row in response.css('tbody tr'):
            evolves_from = row.css('td:nth-child(1) a::attr(title)').get().replace('View Pokedex for #','')[:4]
            evolves_from_form = row.css('td:nth-child(1) small.text-muted::text').get()
            evolves_to = row.css('td:nth-child(3) a::attr(title)').get().replace('View Pokedex for #','')[:4]
            evolves_to_form = row.css('td:nth-child(3) small.text-muted::text').get()
            method = response.meta['page_type']

            if method == 'stone':
                item_used = row.css('td.cell-med-text').xpath('string()').get().strip().replace('outside Alola','')
                item_held = ''
                other_condition = ''
            elif method == 'trade':
                item_used = ''
                item_held = row.css('td.cell-med-text').xpath('string()').get().strip().replace('Kings Rock',"King's Rock")
                other_condition = ''
            else:
                item_used = ''
                item_held = ''
                other_condition = row.css('td.cell-med-text').xpath('string()').get().strip().replace('outside Alola','').replace('or Level 17 in Legends: Arceus','')

            def regional_forms_check(var, var_form):
                '''Checks if the evolution involves a regional form and adds appropriate suffixes to Pokedex numbers while rewriting the form for easier filtering'''
                if var_form is not None:
                    if var_form[0:6] == 'Alolan':
                        var += 'a'
                        var_form = 'Alolan Form'
                    elif var_form[0:8] == 'Galarian':
                        var += 'g'
                        var_form = 'Galarian Form'
                    elif var_form[0:7] == 'Hisuian':
                        var += 'h'
                        var_form = 'Hisuian Form'
                    elif var_form[0:7] == 'Paldean':
                        var += 'p'
                        var_form = 'Paldean Form'
                    else:
                        var += var_form[0:2].lower()
                        var_form = var_form
                
                return var, var_form
            
            pk, pkf = regional_forms_check(evolves_from, evolves_from_form)
            et, etf = regional_forms_check(evolves_to, evolves_to_form)
            
            entry = {
                'evolves_from': pk,
                'evolves_from_form': pkf,
                'evolves_to': et,
                'evolves_to_form': etf,
                'method': method,
                'level': row.css('td.cell-num::text').get(),
                'item_used': item_used,
                'item_held': item_held,
                'stat': '',
                'time_of_day': '',
                'location': '',
                'move_learned': '',
                'gender': '',
                'in_party': '',
                'weather': '',
                'versions': '',
                'other_condition': other_condition,
                'notes': ''
            }
        
            evo_data.append(entry)

    def closed(self, data):
        '''Convert data to pandas DataFrame, clean and filter per gen, and save to Excel file'''

        # create datafrane from the global list 'evo_data'
        df = pd.DataFrame(evo_data)

        # pre-sort cleaning, accounting for whether or not different forms need to appear as separate entries, multiple evolution methods
        # Burmy's forms only differ in how they evolve, but forms are mutable within the space of a battle
        df.loc[df['evolves_from'].str.startswith('0412'), 'evolves_from'] = '0412'
        df.loc[df['evolves_to'].str.startswith('0413'), 'notes'] = "Wormadam will take on Burmy's form at time of evolution, which can change based on the location of its most recent battle, but only if it actually leaves the ball."

        # Darmanitan's standard form doesn't need a suffix, and having it makes the Galarian form weird
        df.loc[df['evolves_to'].str.startswith('0555'), 'evolves_to'] = '0555'
        df.loc[df['evolves_from'] == '0554g', 'evolves_to'] = '0555g'

        # Add rows for Pokemon with multiple evolution methods, or methods that differ between versions
        df.loc[len(df)] = ['0064','','0065','','stone','','Linking Cord','','','','','','','','','Legends: Arceus','','']
        df.loc[len(df)+1] = ['0067','','0068','','stone','','Linking Cord','','','','','','','','','Legends: Arceus','','']
        df.loc[len(df)+2] = ['0075','','0076','','stone','','Linking Cord','','','','','','','','','Legends: Arceus','','']
        df.loc[len(df)+3] = ['0082','','0462','','level','','','','','','Mt Coronet (any section)','','','','','Brilliant Diamond, Shining Pearl','','']
        df.loc[len(df)+4] = ['0082','','0462','','level','','','','','','Coronet Highlands (any section)','','','','','Legends: Arceus','','']
        df.loc[len(df)+5] = ['0093','','0094','','stone','','Linking Cord','','','','','','','','','Legends: Arceus','','']
        df.loc[len(df)+6] = ['0133','','0470','','level','','','','','','Eterna Forest (southwest)','','','','','Brilliant Diamond, Shining Pearl','','']
        df.loc[len(df)+7] = ['0133','','0470','','level','','','','','','The Heartwood (west)','','','','','Legends: Arceus','','']
        df.loc[len(df)+8] = ['0133','','0471','','level','','','','','','Route 217 (northeast)','','','','','Brilliant Diamond, Shining Pearl','','']
        df.loc[len(df)+9] = ['0133','','0471','','level','','','','','','Icepeak Cavern','','','','','Legends: Arceus','','']
        df.loc[len(df)+10] = ['0155','','0156h','','level','17','','','','','','','','','','Legends: Arceus','','']
        df.loc[len(df)+11] = ['0299','','0476','','level','','','','','','Mt Coronet (any section)','','','','','Brilliant Diamond, Shining Pearl','','']
        df.loc[len(df)+12] = ['0299','','0476','','level','','','','','','Coronet Highlands (any section)','','','','','Legends: Arceus','','']
        df.loc[len(df)+13] = ['0854','','0855','','stone','','Chipped Pot','','','','','','','','','Sword, Shield','Antique form','']

        # Pumpkaboo's forms aren't all present in the evolist, but it is necessary to specify each evolution
        pumpkaboo_small = df.loc[df['evolves_from'] == '0710av'].copy()
        pumpkaboo_small.loc[:, ['evolves_from','evolves_to']] = ['0710sm','0711sm']
        pumpkaboo_large = df.loc[df['evolves_from'] == '0710av'].copy()
        pumpkaboo_large.loc[:, ['evolves_from','evolves_to']] = ['0710la','0711la']
        pumpkaboo_super = df.loc[df['evolves_from'] == '0710av'].copy()
        pumpkaboo_super.loc[:, ['evolves_from','evolves_to']] = ['0710su','0711su']
        df = pd.concat([df, pumpkaboo_small, pumpkaboo_large, pumpkaboo_super], ignore_index=True)

        # Lycanroc's Midday and Midnight forms need different abbreviations, and Own Tempo Rockruff isn't a different form outside of how it is acquired
        df.loc[df['evolves_to_form'] == 'Midday Form', 'evolves_to'] = '0745md'
        df.loc[df['evolves_to_form'] == 'Midnight Form', 'evolves_to'] = '0745mn'
        df.loc[df['evolves_to_form'] == 'Dusk Form', ['evolves_from','evolves_from_form','time_of_day','other_condition']] = ['0744','','Dusk','Ability Own Tempo']

        # sort by versions, then method, then evolves_to, then evolves_from
        df['method_index'] = df['method'].apply(lambda x: methods.index(x))
        df['games_index'] = df['versions'].apply(lambda x: games.index(x))
        df = df.sort_values(by=['evolves_from','evolves_to','method_index','games_index'])
        df = df.drop(['method_index','games_index'], axis=1)

        # remove duplicate entries
        duplicate_mask = df.duplicated(subset=['evolves_from','evolves_to','versions'], keep='first')
        df = df[~duplicate_mask]

        # post-sort cleaning, moving evolution conditions into the appropriate columns
        def condition_mover(search_str,correct_col,correct_str):
            '''Locates rows containing a given string in item_used or other_condition (or both), removes it from those columns, and adds a replacement string in the target column.'''
            df[correct_col] = df.apply(lambda row: correct_str if (search_str in row['item_used'] or search_str in row['other_condition']) else row[correct_col], axis=1)
            df['item_used'] = df['item_used'].str.replace(search_str, '')
            df['other_condition'] = df['other_condition'].str.replace(search_str, '')
        
        # evolution cases requiring special handling before condition_mover() is run
        df.loc[df['evolves_to'] == '0350', ['item_held','notes']] = ['Prism Scale','Can still evolve by leveling a Feebas imported from earlier gen with max Beauty, despite being invisible']
        df.loc[df['evolves_to'] == '0892ra', ['method','item_used','other_condition']] = ['status','','Complete the Tower of Water']
        df.loc[df['evolves_to'] == '0892si', ['method','item_used','other_condition']] = ['status','','Complete the Tower of Darkness']

        # evolutions with conditions in the wrong column
        condition_mover('Daytime','time_of_day','Daytime')
        condition_mover('Nighttime','time_of_day','Nighttime')
        condition_mover('in Galar','versions','Sword, Shield')
        condition_mover('outside Galar','location','Outside Galar')
        condition_mover('in Legends: Arceus','versions','Legends: Arceus')
        condition_mover('in buildings','location','In buildings')
        condition_mover('in caves','location','In caves')
        condition_mover('in grass','location','In grass')
        condition_mover('Female','gender','Female')
        condition_mover('Male','gender','Male')
        condition_mover('during rain','weather','Rain')

        # evolution cases requiring special handling after condition_mover() is run
        df.loc[df['evolves_to'] == '0156', 'versions'] = 'Sword, Shield, Brilliant Diamond, Shining Pearl'
        df.loc[df['evolves_to'] == '0589', ['item_held','other_condition']] = ['','Trade with Shelmet']
        df.loc[df['evolves_to'] == '0617', ['item_held','other_condition']] = ['','Trade with Karrablast']
        df.loc[df['evolves_to'] == '0687', 'other_condition'] = 'Hold console upside down (handheld mode only)'
        df.loc[df['evolves_to'] == '0700', ['stat','other_condition']] = ['Friendship','after Fairy-type move learned']
        df.loc[df['evolves_to'] == '0791', ['versions','other_condition']] = ['Sword','']
        df.loc[df['evolves_to'] == '0792', ['versions','other_condition']] = ['Shield','']
        df.loc[df['evolves_to'] == '0809', ['versions','other_condition']] = ['Pokémon GO','400 Meltan Candies']
        df.loc[df['evolves_to'] == '0849am', 'other_condition'] = 'Adamant, Brave, Docile, Hardy, Hasty, Impish, Jolly, Lax, Naive, Naughty, Rash, Quirky, Sassy Nature'
        df.loc[df['evolves_to'] == '0849lo', 'other_condition'] = 'Bashful, Bold, Calm, Careful, Gentle, Lonely, Mild, Modest, Quiet, Relaxed, Serious, Timid Nature'
        df.loc[(df['evolves_to'] == '0855') & (df['item_used'] == 'Cracked Pot'), 'other_condition'] = 'Phony form'
        df.loc[df['evolves_to'] == '0855', 'versions'] = ''
        df.loc[df['evolves_to'] == '0867', ['location','versions','other_condition']] = ['Dusty Bowl','Sword, Shield','Take 49+ damage taken in one battle, then walk under stone sculpture']
        df.loc[df['evolves_to'] == '0869', 'other_condition'] = 'Spin around holding a Sweet item'
        df.loc[df['evolves_to'] == '0899', 'versions'] = 'Legends: Arceus'
        df.loc[df['evolves_to'] == '0901', ['item_used','weather','versions','notes']] = ['Peat Block','Full moon','Legends: Arceus',"The moon doesn't need to be visible"]
        df.loc[df['evolves_to'] == '0902fe', ['method','other_condition']] = ['level','Receive 294 recoil damage in battle without fainting']
        df.loc[df['evolves_to'] == '0902ma', ['method','other_condition']] = ['level','Receive 294 recoil damage in battle without fainting']
        df.loc[df['evolves_to'] == '0903', ['method','item_used','item_held']] = ['level','','Razor Claw']
        df.loc[df['evolves_to'] == '0904', 'versions'] = 'Legends: Arceus'
        df.loc[df['location'] == 'Outside Galar', ['location','versions']] = ['','Brilliant Diamond, Shining Pearl, Legends: Arceus']
        df.loc[df['item_used'].str.endswith('Magnetic Field area'), ['method','item_used','versions']] = ['item','Thunder Stone','Sword, Shield']
        df.loc[df['item_used'].str.endswith('Icy Rock'), ['method','item_used','versions']] = ['item','Ice Stone','Sword, Shield']
        df.loc[df['item_used'].str.endswith('Mossy Rock'), ['method','item_used','versions']] = ['item','Leaf Stone','Sword, Shield']
        df.loc[df['evolves_to'] == '0738', 'versions'] = ''

        # remove leading and trailing commas and spaces
        df.loc[df['item_used'].str.endswith(', '), 'item_used'] = df['item_used'].str.rstrip(', ')
        df.loc[df['item_used'].str.endswith(','), 'item_used'] = df['item_used'].str.rstrip(',')
        df.loc[df['other_condition'].str.startswith(', '), 'other_condition'] = df['other_condition'].str[2:]
        df.loc[df['other_condition'].str.endswith(', '), 'other_condition'] = df['other_condition'].str.rstrip(', ')

        # move from other_condition to move_learned
        move_condition = df['other_condition'].str.endswith('learned')
        df.loc[move_condition, 'move_learned'] = df.loc[move_condition, 'other_condition'].str[6:-8]
        df.loc[move_condition, 'method'] = 'level'
        df.loc[move_condition, 'other_condition'] = ''

        # move from other_condition to item_held
        item_condition = df['other_condition'].str.startswith('hold ')
        df.loc[item_condition, 'item_held'] = df.loc[item_condition, 'other_condition'].str[5:]
        df.loc[item_condition, 'method'] = 'level'
        df.loc[item_condition, 'other_condition'] = ''

        # move from other_condition to stat
        stat_condition = df['other_condition'].str.startswith('Attack')
        df.loc[stat_condition, 'stat'] = df.loc[stat_condition, 'other_condition'].str[:]
        df.loc[stat_condition, 'other_condition'] = ''

        # move from other_condition to in_party
        df.loc[df['evolves_to'] == '0292', ['in_party','other_condition']] = ['Empty spot','Pokéball in bag']
        df.loc[df['evolves_to'] == '0226', ['method','in_party','other_condition']] = ['level','Remoraid','']
        df.loc[df['evolves_to'] == '0675', ['in_party','other_condition']] = ['Dark-type Pokémon','']

        # pokemon available in this generation, but who cannot be evolved in it
        for i in ['0026a','0103a','0105a']:
            df.loc[df['evolves_to'] == i, ['method','level','item_used','other_condition']] = ['N/A','','','Must be evolved in other generation and traded to SwSh']

        # methods need to be renamed for better accuracy
        df.loc[df['method'] == 'stone', 'method'] = 'item'
        df.loc[df['method'] == 'status', 'method'] = 'miscellaneous'
        df.loc[df['method'] == 'friendship', ['method','stat']] = ['level','Friendship']

        # convert level to int, empty cells to NaN
        df['level'] = pd.to_numeric(df['level'], errors='coerce')
        df = df.replace('', pd.NA)

        # capitalize the first letter of other_condition
        df['other_condition'] = df['other_condition'].str[0].str.upper() + df['other_condition'].str[1:]

        # filter out entries not in the associated generation
        df = df[df['evolves_from'].between('0001', '0905') & ~df['evolves_from'].str.match(r'\d{4}[p]$')]
        df = df[df['evolves_to'].between('0001', '0905') & ~df['evolves_to'].str.match(r'\d{4}[p]$')]

        # remove dexited mons, including Nosepass/Probopass in SwSh
        nosepass_mag_field = df[(df['evolves_to'] == '0476') & (df['method'] == 'miscellaneous')].index
        df = df.drop(nosepass_mag_field)
        df = df[~df['evolves_to'].isin(unavailable_mons)]

        # drop evolves_from_form, evolves_to_form, and any empty columns
        df = df.drop(['evolves_from_form','evolves_to_form'], axis=1)
        df = df.dropna(axis=1,how='all')

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g8_evo.xlsx')
        df.to_sql('evo', engine, index=False)
        #df.to_csv('pkmn_g8_evo.csv', index=False)
