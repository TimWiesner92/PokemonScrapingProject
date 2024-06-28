import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

df = pd.DataFrame(columns=['dex_no','move','level','tm_hm','egg_moves','crystal_move_tutor','g1_tm_hm'])
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g2')

class PkmnSpider(scrapy.Spider):
    name = 'pkmn_g2_learnsets'
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex-gs/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]')[0:2].css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_learnset_page)

    def parse_learnset_page(self, response):
        dex_no = response.css('td:not([class])::text')[11].get().replace('#','').zfill(4)
        moves1 = response.xpath('//table[.//tr[1]/td//text()="Generation II Level Up"]').xpath('.//tr[position()>2]').css('a::text').getall()
        levels = [int(s) if s.isdigit() else s for s in response.xpath('//table[.//tr[1]/td//text()="Generation II Level Up"]').xpath('.//tr[position() > 2 and position() mod 2 = 1]').css('td.fooinfo::text').getall()]
        moves2 = response.xpath('//table[.//tr[1]/td//text()="TM & HM Attacks"]').xpath('.//tr[position() > 2]').css('a::text').getall()
        tmhm = response.xpath('//table[.//tr[1]/td//text()="TM & HM Attacks"]').xpath('.//tr[position() > 2 and position() mod 2 = 1]').xpath('.//td[@class="fooinfo" and position() mod 2 = 1]/text()').getall()
        moves3 = response.xpath('//table[.//tr[1]/td[contains(text(), "Egg Moves")]]').xpath('.//tr[position() > 2]').css('a::text').getall()
        moves4 = response.xpath('//table[.//tr[1]/td[contains(text(), "Crystal Move Tutor Attacks")]]').xpath('.//tr[position() > 2]').css('a::text').getall()
        moves5 = response.xpath('//table[.//tr[1]/td[contains(text(), "Gen I Only Moves")]]').xpath('.//tr[position() > 2]').css('a::text').getall()
        g1tmhm = response.xpath('//table[.//tr[1]/td[contains(text(), "Gen I Only Moves")]]').xpath('.//tr[position() > 2]').xpath('.//td[@class="fooinfo" and @colspan = 2]/text()').getall()

        dict1 = {'dex_no': [dex_no] * len(moves1), 'move': moves1, 'level': levels}
        df1 = pd.DataFrame(dict1)

        dict2 = {'dex_no': [dex_no] * len(moves2), 'move': moves2, 'tm_hm': tmhm}
        df2 = pd.DataFrame(dict2)

        dict3 = {'dex_no': [dex_no] * len(moves3), 'move': moves3, 'egg_moves': 'Egg'}
        df3 = pd.DataFrame(dict3)

        dict4 = {'dex_no': [dex_no] * len(moves4), 'move': moves4, 'crystal_move_tutor': '4000 Coins'}
        df4 = pd.DataFrame(dict4)

        dict5 = {'dex_no': [dex_no] * len(moves5), 'move': moves5, 'g1_tm_hm': g1tmhm}
        df5 = pd.DataFrame(dict5)

        dfn1 = pd.merge(df1, df2, on=['dex_no','move'], how='outer').sort_values(by=['move'])

        dfn2 = pd.merge(dfn1, df3, on=['dex_no','move'], how='outer').sort_values(by=['move'])

        dfn3 = pd.merge(dfn2, df4, on=['dex_no','move'], how='outer').sort_values(by=['move'])

        dfn4 = pd.merge(dfn3, df5, on=['dex_no','move'], how='outer').sort_values(by=['move'])

        global df
        df = pd.concat([df, dfn4], ignore_index = True)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''
        global df
        df = df.sort_values(by=['dex_no','level','tm_hm','move'])

        # convert level to int, empty cells to NaN
        df['level'] = pd.to_numeric(df['level'], errors='coerce')

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g2_learnsets.xlsx')
        df.to_sql('learnsets', engine, index=False)
        #df.to_csv('pkmn_g2_learnsets.csv', index=False)
