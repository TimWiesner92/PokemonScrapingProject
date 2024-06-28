import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

move_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g1')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g1_moves"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/attackdex-rby/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]').css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        s = response.css('td.cen::text')[6].get().replace('\r\n\t\t','').replace('\t','').replace('--','—').replace(' ','').replace('%','')
        if s.isdigit():
            effect_rate = int(s)
        else:
            effect_rate = s

        entry = {
            'move_name': response.css('td.cen::text')[0].get().replace('\r\n\t\t',''),
            'move_type': response.css('td.cen')[1].css('img').xpath('@src').getall()[0].title().replace('/Pokedex-Bw/Type/','').replace('.Gif',''),
            'power_points': int(response.css('td.cen::text')[3].get().replace('\r\n\t\t','').replace('\t','')),
            'base_power': int(response.css('td.cen::text')[4].get().replace('\r\n\t\t','').replace('\t','')),
            'accuracy': float(response.css('td.cen::text')[5].get().replace('\r\n\t\t','').replace('\t','')),
            'battle_effect': response.css('td.fooinfo::text')[0].get().replace('\r\n\t\t','').replace('\t',''),
            'secondary_effect': response.css('td.fooinfo::text')[1].get().replace('\r\n\t\t','').replace('\t',''),
            'effect_rate': effect_rate,
            'tm_no': response.css('td.cen::text')[7].get().replace('\r\n\t\t','').replace('\t','').replace('None',''),
            'speed_priority': int(response.css('td.cen::text')[8].get().replace('\r\n\t\t','').replace('\t','')),
            'hit_range': response.css('td.cen::text')[9].get().replace('\r\n\t\t','').replace('\t','')
        }

        move_data.append(entry)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''

        # create datafrane from the global list 'move_data'
        df = pd.DataFrame(move_data)
        df = df.sort_values(by=['move_name'])

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g1_moves.xlsx')
        df.to_sql('moves', engine, index=False)
        #df.to_csv('pkmn_g1_moves.csv', index=False)
