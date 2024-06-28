import scrapy, pandas as pd, math
from openpyxl import load_workbook
from sqlalchemy import create_engine

pk_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g3')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g3_pkmn"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex-rs/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]')[0:3].css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        def reduce_fractions(num, denom):
            '''Function to reduce a fraction using greatest common demoninator. E.g., reduce '5/25' down to '1/5'. Takes num and denom, reduces each, and returns reduced_num and reduced_denom'''
            gcd = math.gcd(int(num), int(denom))
            reduced_num = int(num) // gcd
            reduced_denom = int(denom) // gcd
            return reduced_num, reduced_denom
        
        # converts raw data on gender into 'N/A' for genderless pokemon, 'Always female' or 'Always male' for those pokemon,
        # or produces a simplified ratio (e.g. '7:1' for startes) for pokemon that can be male or female
        if response.xpath('//td[b/text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().strip() == 'Genderless':
            mfr = 'N/A'
        else:
            mr = response.xpath('//td[b/text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().strip().replace('Male: ','').replace(' %','')
            fr = response.xpath('//td[b/text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().strip().replace('Female: ','').replace(' %','')
            if mr == '0' and fr == '0':
                mfr = 'N/A'
            elif mr == '0':
                mfr = 'Always female'
            elif fr == '0':
                mfr = 'Always male'
            else:
                reduced_mr, reduced_fr = reduce_fractions(int(100 * float(mr)), int(100 * float(fr)))
                mfr = str(reduced_mr) + ':' + str(reduced_fr)

        # splits awarded EVs for beating a pokemon into individual values so they each can be placed in their own columns for easier sorting and filtering
        hp = 0
        atk = 0
        dfn = 0
        spa = 0
        spd = 0
        spe = 0
        for i in response.xpath('//td[text()="Egg Steps to Hatch"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().strip().replace('HP ','HP: ').replace('Attack ','Attack: ').replace('Defense ','Defense: ').replace('Special Attack ','Special Attack: ').replace('Special Defense ','Special Defense: ').replace('Speed ','Speed: ').split(', '):
            if i[0:2] == 'HP':
                hp = int(i[4])
            elif i[0:6] == 'Attack':
                atk = int(i[8])
            elif i[0:7] == 'Defense':
                dfn = int(i[9])
            elif i[0:14] == 'Special Attack':
                spa = int(i[16])
            elif i[0:15] == 'Special Defense':
                spd = int(i[17])
            else:
                spe = int(i[7])

        # height, abilities, and egg_groups declared here for simpler list comprehensions
        # splits height into feet and inches so height can be transformed into a float
        height = response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[4]/text()').get().strip().replace('"','').split('\'')
        # splits abilities into individual values to put in separate columns for easier filtering
        abilities = response.xpath('//td[@colspan="4"]//b/text()').get().strip().replace('Ability: ','').split(' & ')
        egg_groups = response.xpath('//table[@class="dextable"][.//td[@class="fooevo" and contains(text(), "Egg Groups")]]').css('td::text').getall()[3:]

        forms_list = response.css('table.dextable').css('td.fooinfo[colspan="7"]').getall()
        forms = [form.replace('<td class="fooinfo" colspan="7"><b>Stats</b>', '').replace('</td>', '') for form in forms_list]
        for form in forms:
            if len(form) == 0:
                dex = response.css('td.fooinfo::text')[3].get().strip().zfill(4)
            else:
                dex = response.css('td.fooinfo::text')[3].get().strip().zfill(4) + form[2:4].lower()
                
            entry = {
                'dex_no': dex,
                'name': response.css('td.fooinfo::text')[5].get().strip().zfill(4),
                'form': form.replace(' (','').replace(')','').replace('Defence','Defense').replace('Form','form'),
                'class': response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().strip().replace(' Pokémon',''),
                'dex_cat': response.xpath('//td[text()="Dex Category"]/ancestor::tr/following-sibling::tr[1]//td[2]//img/@title').get().strip().capitalize(),
                'color': response.xpath('//td[text()="Dex Category"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().strip().capitalize(),
                'height_ft': round((float(height[0]) + (float(height[1]) / 12)),2),
                'weight_lb': float(response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[5]/text()').get().strip().replace(' lbs','').replace(',','')),
                'male_female_rate': mfr,
                'type_one': response.xpath('//table[@class="dextable"]//img[starts-with(@src, "/pokedex-rs/type/")]/@src').getall()[0].replace('/pokedex-rs/type/','').replace('.gif','').capitalize(),
                'type_two': response.xpath('//table[@class="dextable"]//img[starts-with(@src, "/pokedex-rs/type/")]/@src').getall()[1].replace('/pokedex-rs/type/na.gif','').replace('/pokedex-rs/type/','').replace('.gif','').capitalize(),
                'ability_one': abilities[0],
                'ability_two': abilities[1] if len(abilities) == 2 else '',
                'base_hp': int(response.xpath('//table[@class="dextable"][.//td[b/text()="Stats" and contains(text(), "' + form + '")]]//td[@class="cen"]/text()').getall()[0]),
                'base_atk': int(response.xpath('//table[@class="dextable"][.//td[b/text()="Stats" and contains(text(), "' + form + '")]]//td[@class="cen"]/text()').getall()[1]),
                'base_def': int(response.xpath('//table[@class="dextable"][.//td[b/text()="Stats" and contains(text(), "' + form + '")]]//td[@class="cen"]/text()').getall()[2]),
                'base_sp_atk': int(response.xpath('//table[@class="dextable"][.//td[b/text()="Stats" and contains(text(), "' + form + '")]]//td[@class="cen"]/text()').getall()[3]),
                'base_sp_def': int(response.xpath('//table[@class="dextable"][.//td[b/text()="Stats" and contains(text(), "' + form + '")]]//td[@class="cen"]/text()').getall()[4]),
                'base_speed': int(response.xpath('//table[@class="dextable"][.//td[b/text()="Stats" and contains(text(), "' + form + '")]]//td[@class="cen"]/text()').getall()[5]),
                'cptr_rt': int(response.xpath('//td[text()="Egg Steps to Hatch"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().strip()),
                'base_egg_steps': int(response.xpath('//td[text()="Egg Steps to Hatch"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().strip().replace(',','').replace(' Steps','')),
                'egg_group_one': egg_groups[0].replace('Cannot Breed','No Eggs Discovered'),
                'egg_group_two': egg_groups[1] if len(egg_groups) == 2 else '',
                'hp_evs': hp,
                'atk_evs': atk,
                'def_evs': dfn,
                'sp_atk_evs': spa,
                'sp_def_evs': spd,
                'speed_evs': spe
            }

            pk_data.append(entry)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''

        # create datafrane from the global list 'pk_data'
        df = pd.DataFrame(pk_data)
        df = df.sort_values(by=['dex_no'])

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g3_pkmn.xlsx')
        df.to_sql('pkmn', engine, index=False)
        #df.to_csv('pkmn_g3_pkmn.csv', index=False)
