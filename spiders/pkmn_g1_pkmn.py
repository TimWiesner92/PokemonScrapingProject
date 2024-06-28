import scrapy, pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

pk_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g1')

class PkmnSpider(scrapy.Spider):
    name = 'pkmn_g1_pkmn'
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]').css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_pkmn_page)

    def parse_pkmn_page(self, response):
        # height declared here for simpler list comprehension
        height = response.css('td.fooinfo::text')[5].get().replace('"','').split('\'')

        entry = {
            'dex_no': response.css('td.fooinfo::text')[3].get().replace('#','').zfill(4),
            'name': response.css('td.fooinfo::text')[2].get(),
            'sprite': response.css('table.dextab').css('img').xpath('@src').get().replace('/pokedex/icon/','').replace('.png','').title(),
            'type_one': [e.replace('/pokedex-bw/type/','').replace('.gif','') for e in response.css('td.cen')[0].css('img').xpath('@src').getall()][0].title(),
            'type_two': [e.replace('/pokedex-bw/type/','').replace('.gif','') for e in response.css('td.cen')[0].css('img').xpath('@src').getall()][1].title() if (len(response.css('td.cen')[0].css('img').xpath('@src').getall()) == 2) else '',
            'class': response.css('td.fooinfo::text')[4].get()[:-8],
            'height_ft': round((float(height[0]) + (float(height[1]) / 12)),2),
            'weight_lb': float(response.css('td.fooinfo::text')[7].get().replace('lbs','')),
            'cptr_rt': int(response.css('td.fooinfo::text')[9].get()),
            'exp_rt': response.css('td.fooinfo::text')[11].get().replace('1,250,000 Points','Slow'),
            'hp': int(response.css('td.fooinfo::text')[-18].get()),
            'atk': int(response.css('td.fooinfo::text')[-17].get()),
            'def': int(response.css('td.fooinfo::text')[-16].get()),
            'special': int(response.css('td.fooinfo::text')[-15].get()),
            'speed': int(response.css('td.fooinfo::text')[-14].get())
        }
        
        pk_data.append(entry)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''

        # create datafrane from the global list 'pk_data'
        df = pd.DataFrame(pk_data)
        df = df.sort_values(by=['dex_no'])

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g1_pkmn.xlsx')
        df.to_sql('pkmn', engine, index=False)
        #df.to_csv('pkmn_g1_pkmn.csv', index=False)
