import scrapy, pandas as pd, math, re
from openpyxl import load_workbook
from sqlalchemy import create_engine

pk_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g5')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g5_pkmn"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex-bw/"]

    def parse(self, response):
        '''Method for extracting links to pages to crawl from the pulldowns'''
        pulldown_links = response.xpath('//select[@name="SelectURL"]')[0:5].css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        '''Method for pre-processing and allocating data pulled from each page'''
        # define functions used in processing data
        def reduce_fractions(num, denom):
            '''Function to reduce a fraction using greatest common demoninator. E.g., reduce '5/25' down to '1/5'. Takes num and denom, reduces each, and returns reduced_num and reduced_denom'''
            gcd = math.gcd(int(num), int(denom))
            reduced_num = int(num) // gcd
            reduced_denom = int(denom) // gcd
            return reduced_num, reduced_denom
        
        def extract_types(html_string, pat1, pat2, pk_name):
            '''Function to extract text from an html string based on a pre-defined pattern'''
            matches = pat1.findall(html_string)
            types_match = [match.capitalize() for match in matches]
            form_match = pat2.search(html_string).group(1)
            # Darmanitan's forms are named differently in Type than they are in Alternate Forms
            if pk_name == 'Darmanitan' and form_match == 'Normal':
                form_match = 'Standard Mode'
            return (form_match,types_match)
                    
        # converts raw data on gender into 'N/A' for genderless pokemon, 'Always female' or 'Always male' for those pokemon,
        # or produces a simplified ratio (e.g. '7:1' for startes) for pokemon that can be male or female
        if response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').get()[-15:-5] == 'Genderless':
            mfr = 'N/A'
        else:
            mr = response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').css('td::text').getall()[2].replace('%','')
            fr = response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').css('td::text').getall()[5].replace('%','')
            if mr == '0' and fr == '0':
                mfr = 'double_zero (set back to N/A if this shows up, remove this clause if not)'
            elif mr == '0':
                mfr = 'Always female'
            elif fr == '0':
                mfr = 'Always male'
            else:
                reduced_mr, reduced_fr = reduce_fractions(int(100 * float(mr)), int(100 * float(fr)))
                mfr = str(reduced_mr) + ':' + str(reduced_fr)

        # splits awarded EVs for beating a pokemon into individual values so they each can be placed in their own columns for easier sorting and filtering
        evs_html = response.xpath('//td[text()="Effort Values Earned"]/ancestor::tr/following-sibling::tr[1]//td[3]').get().replace('<td class="fooinfo">','').replace('</td>','')[:-4].split('<br><b>')
        def ev_split(html):
            def list_split(ev_list):
                evs = [0,0,0,0,0,0]
                for i in ev_list:
                    if i[2:4] == 'HP':
                        evs[0] = int(i[0])
                    elif i[2:8] == 'Attack':
                        evs[1] = int(i[0])
                    elif i[2:9] == 'Defense':
                        evs[2] = int(i[0])
                    elif i[2:12] == 'Sp. Attack':
                        evs[3] = int(i[0])
                    elif i[2:13] == 'Sp. Defense':
                        evs[4] = int(i[0])
                    else:
                        evs[5] = int(i[0])
                return evs

            if len(html) == 1:
                html_split = html[0].split('<br>')
                ev_iter = list_split(html_split)
            else:
                forms = []
                points = []
                for i in html:
                    form, point = i.split('</b><br>')
                    forms.append(form)
                    points.append(point)
                forms = [i.replace('<b>','') for i in forms]
                points = [i.split('<br>') for i in points]
                ev_list_split = [list_split(i) for i in points]
                ev_iter = {forms[i]: ev_list_split[i] for i in range(len(html))}
            return ev_iter
        
        # height, abilities, and egg_groups declared here for simpler list comprehensions
        # splits height into feet and inches so height can be transformed into a float
        height = response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().replace('”','').split('’')
        # splits abilities into individual values to put in separate columns for easier filtering
        abilities = ''.join(response.xpath('//td[b/text()="Abilities"]//text()').getall()).strip().replace('Abilities: ','').replace('- B2W2','B2W2').split(' - ')
        if response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"][2]//text()').get()[-12:] == 'cannot breed':
            eg1,eg2 = ['No Eggs Discovered','']
        elif len(response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"][2]//a//text()').getall()) == 1:
            eg1,eg2 = [response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"][2]//a//text()').getall()[0],'']
        else:
            eg1 = response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"][2]//a//text()').getall()[0]
            eg2 = response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"][2]//a//text()').getall()[1]
        
        # set up types_html, stats_forms, and name for if/else block
        types_html = response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//tr').getall()
        stats_forms = response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])//td[@class="fooinfo" and text()="Base Stats"]').getall()
        name = response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get()

        # set up forms and types to be parsed based on whether a pokemon has multiple distinct forms
        if len(response.xpath('//td[b[contains(text(), "Stats")]]').getall()) > 1:
            forms = response.xpath('//td[text()="Alternate Forms"]/ancestor::tr/following-sibling::tr[1]//td/b/text()').getall()
            if len(types_html) != 0:
                p1 = re.compile(r'/pokedex-dp/type/([a-z]+)\.gif')
                p2 = re.compile(r'<td width="50%">(.*?)</td>')
                types = {form_match: types_match for form_match, types_match in (extract_types(html,p1,p2,name) for html in types_html)}
            else:
                types = [i.replace('/pokedex-dp/type/','').replace('.gif','').capitalize() for i in response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//img[starts-with(@src, "/pokedex-dp/type/")]/@src').getall()]
        else:
            forms = ['']
            types = [i.replace('/pokedex-dp/type/','').replace('.gif','').capitalize() for i in response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//img[starts-with(@src, "/pokedex-dp/type/")]/@src').getall()]

        for form in forms:
            # set up selectors for stats and types based on how a given pokemon's forms vary from one another
            if len(types_html) == len(forms) == len(stats_forms):
                type_sel = types[form]
                stat_sel = forms.index(form) + 1
            elif len(types_html) == 0 and len(forms) == len(stats_forms):
                type_sel = types
                stat_sel = forms.index(form) + 1
            elif len(types_html) == 0 and len(forms) != len(stats_forms):
                type_sel = types
                if forms.index(form) == 0:
                    stat_sel = 1
                else:
                    stat_sel = 2
            elif len(types_html) != len(forms) == len(stats_forms):
                type_sel = types[forms.index(form)]
                if forms.index(form) == 0:
                    stat_sel = 1
                else:
                    stat_sel = 2
            else:
                type_sel = types[form]
                if forms.index(form) == 0:
                    stat_sel = 1
                else:
                    stat_sel = 2
            
            # set up selectors for evs:
            if len(evs_html) == 1:
                evs_sel = ev_split(evs_html)
            else:
                if name == 'Darmanitan' and form == 'Standard Mode':
                    evs_sel = ev_split(evs_html)['Standard']
                elif name == 'Watchog':
                    evs_sel = ev_split(evs_html)['Black 2 &amp; White 2']
                else:
                    evs_sel = ev_split(evs_html)[form]
            
            entry = {
                'dex_no': response.xpath('//td[b[text()="National"]]/following-sibling::td[1]/text()').get().replace('#','').zfill(4),
                'form_index': forms.index(form),
                'name': name,
                'form': form,
                'class': response.xpath('//td[text()="Classification"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().replace(' Pokémon',''),
                'height_ft': round((float(height[0]) + (float(height[1]) / 12)),2),
                'weight_lb': float(response.xpath('//td[text()="Weight"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().replace(',','').replace('lbs','')),
                'male_female_rate': mfr,
                'type_one': type_sel[0],
                'type_two': type_sel[1] if len(type_sel) == 2 else '',
                'ability_one': abilities,
#                'ability_two': abilities[1] if len(abilities) == 2 else '',
                'base_hp': int(response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])[' + str(stat_sel) + ']//td[@class="fooinfo" and text()="Base Stats"]/following-sibling::td[1]/text()').get()),
                'base_atk': int(response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])[' + str(stat_sel) + ']//td[@class="fooinfo" and text()="Base Stats"]/following-sibling::td[2]/text()').get()),
                'base_def': int(response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])[' + str(stat_sel) + ']//td[@class="fooinfo" and text()="Base Stats"]/following-sibling::td[3]/text()').get()),
                'base_sp_atk': int(response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])[' + str(stat_sel) + ']//td[@class="fooinfo" and text()="Base Stats"]/following-sibling::td[4]/text()').get()),
                'base_sp_def': int(response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])[' + str(stat_sel) + ']//td[@class="fooinfo" and text()="Base Stats"]/following-sibling::td[5]/text()').get()),
                'base_speed': int(response.xpath('(//table[@class="dextable"][.//td[b[starts-with(text(), "Stats")]]])[' + str(stat_sel) + ']//td[@class="fooinfo" and text()="Base Stats"]/following-sibling::td[6]/text()').get()),
                'cptr_rt': int(response.xpath('//td[text()="Base Egg Steps"]/ancestor::tr/following-sibling::tr[1]//td[4]/text()').get()),
                'exp_growth': response.xpath('//td[text()="Experience Growth"]/ancestor::tr/following-sibling::tr[1]//td[1]/br/following-sibling::text()').get(),
                'base_egg_steps': int(response.xpath('//td[text()="Base Egg Steps"]/ancestor::tr/following-sibling::tr[1]//td[5]/text()').get().replace(',','')),
                'egg_group_one': eg1,
                'egg_group_two': eg2,
                'hp_evs': evs_sel[0],
                'atk_evs': evs_sel[1],
                'def_evs': evs_sel[2],
                'sp_atk_evs': evs_sel[3],
                'sp_def_evs': evs_sel[4],
                'speed_evs': evs_sel[5]
            }
        
            pk_data.append(entry)
    
    def closed(self, data):
        '''Method for post-processing and saving the data after the spider finishes'''
        # define functions used in this method
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        # create datafrane from the global list 'pk_data'
        df = pd.DataFrame(pk_data)
        df = df.sort_values(by=['dex_no','form_index'])

        # add suffixes to pokemon with alternate forms
        df['dex_no'] = df['dex_no'] + df['form'].str[:2].str.lower()

        # remove suffixes from Pokemon with 'standard' forms
        df.loc[df['dex_no'].str.startswith('0386no'), 'dex_no'] = '0386'
        df.loc[df['dex_no'].str.startswith('0479ro'), 'dex_no'] = '0479'
        df.loc[df['dex_no'].str.startswith('0555st'), 'dex_no'] = '0555'
        df.loc[df['dex_no'].str.startswith('0646ky'), 'dex_no'] = '0646'

        # drop form_index
        df = df.drop(['form_index'], axis=1)

        # save to xlsx, PostgreSQL, or csv
        save_xlsx('pkmn_g5_pkmn.xlsx')
        df.to_sql('pkmn', engine, index=False)
        #df.to_csv('pkmn_g5_pkmn.csv', index=False)
