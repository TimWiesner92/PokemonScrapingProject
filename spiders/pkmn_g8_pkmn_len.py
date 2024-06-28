import scrapy, pandas as pd, math, re
from openpyxl import load_workbook
from sqlalchemy import create_engine

pk_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g8')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g8_pkmn_len"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex-swsh/"]

    def parse(self, response):
        '''Method for extracting links to pages to crawl from the pulldowns'''
        pulldown_links = [x for x in response.xpath('//select[@name="SelectURL"]')[0:8].css('option').xpath('@value').getall() if x != '#']

        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        '''Method for pre-processing and allocating data pulled from each page'''
        # define functions used in processing data
        def reduce_fractions(num, denom):
            '''Function to reduce a fraction using greatest common demoninator. E.g., reduce '5/25' down to '1/5'. Takes num and denom, reduces each, and returns reduced_num and reduced_denom'''
            gcd = math.gcd(int(num), int(denom))
            reduced_num = int(num) // gcd
            reduced_denom = int(denom) // gcd
            return reduced_num, reduced_denom
        
        def extract_types(html_string, pat1, pat2, pk_name):
            '''Function to extract text from an html string based on a pre-defined pattern'''
            matches = pat1.findall(html_string)
            types_match = [match.capitalize() for match in matches]
            form_match = pat2.search(html_string).group(1)
            # Darmanitan's forms are named differently in Type than they are in Alternate Forms
            if pk_name == 'Darmanitan':
                if form_match == 'Normal':
                    form_match = 'Normal Standard Mode'
                elif form_match == 'Zen Mode':
                    if types_match == ['Fire', 'Psychic']:
                        form_match = 'Normal Zen Mode'
                    else:
                        form_match = 'Galarian Form Zen Mode'
                else:
                    form_match = 'Galarian Form Standard Mode'
            elif pk_name == 'Necrozma' and form_match == 'Normal':
                form_match = 'Necrozma'
            elif pk_name == 'Decidueye' and form_match == 'Normal':
                form_match = 'Alolan Form'
            elif form_match == 'Alolan':
                form_match = 'Alola Form'
            elif form_match == 'Galarian':
                form_match = 'Galarian Form'
            elif form_match == 'Hisuian':
                form_match = 'Hisuian Form'
            return (form_match,types_match)
        
        def ev_split(html):
            def list_split(ev_list):
                evs = [0,0,0,0,0,0]
                for i in ev_list:
                    if i[2:4] == 'HP':
                        evs[0] = int(i[0])
                    elif i[2:8] == 'Attack':
                        evs[1] = int(i[0])
                    elif i[2:9] == 'Defense':
                        evs[2] = int(i[0])
                    elif i[2:12] == 'Sp. Attack':
                        evs[3] = int(i[0])
                    elif i[2:13] == 'Sp. Defense':
                        evs[4] = int(i[0])
                    else:
                        evs[5] = int(i[0])
                return evs

            if len(html) == 1:
                html_split = html[0].split('<br>')
                ev_iter = list_split(html_split)
            else:
                forms = []
                points = []
                for i in html:
                    form, point = i.split('</b><br>')
                    forms.append(form)
                    points.append(point)
                forms = [i.replace('<b>','') for i in forms]
                points = [i.split('<br>') for i in points]
                ev_list_split = [list_split(i) for i in points]
                ev_iter = {forms[i]: ev_list_split[i] for i in range(len(html))}
            return ev_iter

        # converts raw data on gender into 'N/A' for genderless pokemon, 'Always female' or 'Always male' for those pokemon,
        # or produces a simplified ratio (e.g. '7:1' for startes) for pokemon that can be male or female
        if response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').get()[-19:-5] == 'Gender Unknown':
            mfr = 'N/A'
        else:
            mr = response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').css('td::text').getall()[2].replace('%','').replace('*','')
            fr = response.xpath('//td[text()="Gender Ratio"]/ancestor::tr/following-sibling::tr[1]//td[4]').css('td::text').getall()[5].replace('%','').replace('*','')
            if mr == '0' and fr == '0':
                mfr = 'double_zero (set back to N/A if this shows up, remove this clause if not)'
            elif mr == '0':
                mfr = 'Always female'
            elif fr == '0':
                mfr = 'Always male'
            else:
                reduced_mr, reduced_fr = reduce_fractions(int(100 * float(mr)), int(100 * float(fr)))
                mfr = str(reduced_mr) + ':' + str(reduced_fr)

        # splits awarded EVs for beating a pokemon into individual values so they each can be placed in their own columns for easier sorting and filtering
        evs_html = response.xpath('//td[text()="Effort Values Earned"]/ancestor::tr/following-sibling::tr[1]//td[3]').get().replace('<td class="fooinfo">','').replace('</td>','')[:-4]
        if 'Zen Mode' in evs_html:
            evs_html = evs_html
        elif 'Alola Form' in evs_html:
            evs_html = 'Normal</b><br>' + evs_html
        elif 'Galarian Form' in evs_html:
            evs_html = 'Normal</b><br>' + evs_html
        elif 'Hisuian Form' in evs_html:
            evs_html = 'Normal</b><br>' + evs_html
        evs_html = evs_html.split('<br><b>')

        # height, abilities, and egg_groups declared here for simpler list comprehensions
        # splits height into feet and inches so height can be transformed into a float
        #base_height_raw = response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().split(' / ')
        #base_height_sliced = [i.replace('"','').replace(' / ','\'').split('\'') for i in base_height_raw]
        #base_height = [round((float(i[0]) + (float(i[1]) / 12)),2) for i in base_height_sliced]
        #base_weight = [float(i) for i in response.xpath('//td[text()="Weight"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().replace(',','').replace('lbs','').split(' / ')]
        # splits abilities into individual values to put in separate columns for easier filtering
        abilities_raw = ''.join(response.xpath('//td[b/text()="Abilities"]//text()').getall()).strip().split(' Abilities: ')
        egg_groups_selector = response.xpath('//td[text()="Egg Groups"]/ancestor::tr/following-sibling::tr[1]//td[@class="fooinfo"]//a//text()').getall()
        egg_steps_selector = response.xpath('//td[text()="Base Egg Steps"]/ancestor::tr/following-sibling::tr[1]//td[5]/text()').get().replace(',','').replace(' (SWSH)','')
            
        # set up types_html, stats_forms, and name for if/else block
        types_html = response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[5]//tr').getall()
        stats_forms = response.xpath('//table[@class="dextable"][.//td[h2[starts-with(text(), "Stats")]]]//td[@class="fooinfo" and contains(text(), "Base Stats")]').getall()
        stats_two = [[int(i) for i in response.xpath('(//table[@class="dextable"][.//td[h2[starts-with(text(), "Stats")]]])[' + str(j + 1) + ']//td[@class="fooinfo" and contains(text(), "Base Stats")]/following-sibling::td/text()').getall()] for j in range(len(stats_forms))]
        name = response.xpath('//td[text()="Type"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get()
        cptr_rt = response.xpath('//td[text()="Capture Rate"]/ancestor::tr/following-sibling::tr[1]//td[4]/text()').get()
        forms = response.xpath('//td[text()="Alternate Forms"]/ancestor::tr/following-sibling::tr[1]//td/b/text()').getall()
        mega_primal_forms = list(set(response.xpath('//td[@class="fooevo" and @colspan="6"]//h2//text()').getall()))

        if len(mega_primal_forms) > 0:
            selector = 'len(mega_primal_forms) > 0'
        else:
            if len(types_html) == len(forms) == len(stats_forms):
                selector = 'len(types_html) == len(forms) == len(stats_forms)'
            elif len(types_html) == len(forms) and len(stats_forms) == 1:
                selector = 'len(types_html) == len(forms) and len(stats_forms) == 1'
            elif len(types_html) == 0 and len(forms) == len(stats_forms):
                selector = 'len(types_html) == 0 and len(forms) == len(stats_forms)'
            elif len(types_html) == 0 and len(forms) != len(stats_forms):
                selector = 'len(types_html) == 0 and len(forms) != len(stats_forms)'
            elif len(types_html) != len(forms) == len(stats_forms):
                selector = 'len(types_html) != len(forms) == len(stats_forms)'
            else:
                selector = 'else'

        entry = {
            'dex_no': response.xpath('//td[b[text()="National"]]/following-sibling::td[1]/text()').get().replace('#','').zfill(4),
            'name': name,
            'selector': selector,
            'mega_primal_forms': len(mega_primal_forms),
            'forms': len(forms),
            'types': len(types_html),
            'stats': len(stats_forms),
            'stats_two': len(stats_two),
            'class': response.xpath('//td[text()="Classification"]/ancestor::tr/following-sibling::tr[1]//td[1]/text()').get().replace(' Pokémon',''),
            'height_ft': len(response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]//text()').getall()),
            'male_female_rate': mfr,
            'ability_one': abilities_raw,
            'cptr_rt': len(cptr_rt),
            'exp_growth': response.xpath('//td[text()="Experience Growth"]/ancestor::tr/following-sibling::tr[1]//td[1]/br/following-sibling::text()').get(),
            'base_egg_steps': 'N/A' if egg_steps_selector == '\xa0' else int(egg_steps_selector),
            'egg_group_one': egg_groups_selector,
            'dynamax': response.xpath('//td[text()="Dynamax Capable?"]/ancestor::tr/following-sibling::tr[1]//td[4]//text()').get().strip()[-11:].replace('can Dynamax','Yes').replace('not Dynamax','No'),
            'evs': evs_html,
        }

        pk_data.append(entry)

    def closed(self, data):
        '''Method for post-processing and saving the data after the spider finishes'''
        # define functions used in this method
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        # create datafrane from the global list 'pk_data'
        df = pd.DataFrame(pk_data)
        df = df.sort_values(by=['dex_no'])

        # save to xlsx, PostgreSQL, or csv
        save_xlsx('pkmn_g8_pkmn_len.xlsx')
        df.to_sql('pkmn_len', engine, index=False)
        #df.to_csv('pkmn_g8_pkmn_len.csv', index=False)
 