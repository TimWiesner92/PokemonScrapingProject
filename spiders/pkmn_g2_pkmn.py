import scrapy, pandas as pd, math
from openpyxl import load_workbook
from sqlalchemy import create_engine

pk_data = []
engine = create_engine('postgresql://postgres:psql_password@localhost:5432/pkmn_g2')

class PkmnSpider(scrapy.Spider):
    name = "pkmn_g2_pkmn"
    allowed_domains = ["www.serebii.net"]
    start_urls = ["https://www.serebii.net/pokedex-gs/"]

    def parse(self, response):
        # Extracting links from the pulldowns
        pulldown_links = response.xpath('//select[@name="SelectURL"]')[0:2].css('option').xpath('@value').getall()
        
        for link in pulldown_links:
            yield scrapy.Request(response.urljoin(link), callback=self.parse_move_page)

    def parse_move_page(self, response):
        def reduce_ratio(num, denom):
            gcd = math.gcd(int(num), int(denom))
            reduced_num = int(num) // gcd
            reduced_denom = int(denom) // gcd
            return reduced_num, reduced_denom
        
        if response.css('td.fooinfo')[4].get()[-15:-5] == 'Genderless':
            mfr = 'N/A'
        else:
            mr = response.css('td.fooinfo')[4].css('td::text')[2].get().replace('%','')
            fr = response.css('td.fooinfo')[4].css('td::text')[5].get().replace('%','')
            if mr == '0':
                mfr = 'Always female'
            elif fr == '0':
                mfr = 'Always male'
            else:
                reduced_mr, reduced_fr = reduce_ratio(int(100 * float(mr)), int(100 * float(fr)))
                mfr = str(reduced_mr) + ':' + str(reduced_fr)
        
        # height declared here for simpler list comprehension
        height = response.xpath('//td[text()="Height"]/ancestor::tr/following-sibling::tr[1]//td[2]/text()').get().replace('"','').split('\'')

        if len(response.css('table.dexitem').css('td').css('a::text').getall()) == 0:
            eg1, eg2 = ['No Eggs Discovered','']
        elif len(response.css('table.dexitem').css('td').css('a::text').getall()) == 1:
            eg1, eg2 = [response.css('table.dexitem').css('td').css('a::text').getall()[0],'']
        else:
            eg1, eg2 = [response.css('table.dexitem').css('td').css('a::text').getall()[0],response.css('table.dexitem').css('td').css('a::text').getall()[1]]
        
        entry = {
            'dex_no': response.css('td:not([class])::text')[11].get().replace('#','').zfill(4),
            'name': response.css('td.fooinfo::text')[0].get(),
            'sprite': response.css('table.dextab').css('img').xpath('@src').get().replace('/pokedex-gs/icon/','').replace('.png','').title(),
            'male_female_rate': mfr,
            'type_one': [e.replace('/pokedex-bw/type/','').replace('.gif','') for e in response.css('td.cen')[0].css('img').xpath('@src').getall()][0].title(),
            'type_two': [e.replace('/pokedex-bw/type/','').replace('.gif','') for e in response.css('td.cen')[0].css('img').xpath('@src').getall()][1].title() if (len(response.css('td.cen')[0].css('img').xpath('@src').getall()) == 2) else '',
            'class': response.css('td.fooinfo')[5].get()[20:-13],
            'height_ft': round((float(height[0]) + (float(height[1]) / 12)),2),
            'weight_lb': float(response.xpath('//td[text()="Weight"]/ancestor::tr/following-sibling::tr[1]//td[3]/text()').get().replace('lbs','')),
            'cptr_rt': int(response.css('td.fooinfo')[8].get()[20:-5]),
            'exp_rt': response.css('td.fooinfo')[10].get().split('<br>')[1][:-5],
            'hp': int(response.css('td.fooinfo::text')[-21].get()),
            'atk': int(response.css('td.fooinfo::text')[-20].get()),
            'def': int(response.css('td.fooinfo::text')[-19].get()),
            'sp_atk': int(response.css('td.fooinfo::text')[-18].get()),
            'sp_def': int(response.css('td.fooinfo::text')[-17].get()),
            'speed': int(response.css('td.fooinfo::text')[-16].get()),
            'base_egg_steps': int(response.css('td.fooinfo')[9].get()[20:-5].replace(',','')),
            'egg_group_one': eg1,
            'egg_group_two': eg2
        }

        pk_data.append(entry)
    
    def closed(self, data):
        '''Method to be called after the spider finishes'''

        # create datafrane from the global list 'pk_data'
        df = pd.DataFrame(pk_data)
        df = df.sort_values(by=['dex_no'])

        # save to xlsx, PostgreSQL, or csv
        def save_xlsx(file_path):
            '''Saves df to excel file, then uses openpyxl to add filters and freeze the top row.'''
            df.to_excel(file_path, index=False)
            # Load the workbook and select the active worksheet
            wb = load_workbook(file_path)
            ws = wb.active

            # Add filters to each column
            ws.auto_filter.ref = ws.dimensions

            # Freeze the top row
            ws.freeze_panes = ws['A2']

            # Save the changes
            wb.save(file_path)

        save_xlsx('pkmn_g2_pkmn.xlsx')
        df.to_sql('pkmn', engine, index=False)
        #df.to_csv('pkmn_g2_pkmn.csv', index=False)
